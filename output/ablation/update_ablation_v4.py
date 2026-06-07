#!/usr/bin/env python3
"""Scan ManiMamba v4 ablation experiment output files and update manimamba-ablation-v4.xlsx.

Tracks 6 variants x 5 datasets:
  - ManiMamba Baseline (02_baseline_v3)
  - alpha+tanh (V4_tanh_alpha)
  - w/o B+C (V4_no_bc)
  - w dt (V4_w_dt)
  - linear interpolation (V4_linear_interp)
  - Geodesic Smoothness Reg (V4_geo_smooth_reg)

Datasets:
  - ETTm1, Weather, ECL, ETTh2  (pred_lens 96/192/336/720)
  - PEMS08                       (pred_lens 12/24/48/96)

Each variant has 4 pred_len rows + 1 Avg row.
Best values are highlighted per (dataset, metric, row_type).

Usage:
    python output/ablation/update_ablation_v4.py [--xlsx PATH] [--output-dir DIR]
    python output/ablation/update_ablation_v4.py --create

Default:
    --xlsx        output/ablation/manimamba-ablation-v4.xlsx
    --output-dir  output/ablation
"""

import argparse
import math
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VARIANTS = [
    ("ManiMamba Baseline", ""),
    ("alpha+tanh", "tanh_alpha"),
    ("w/o B+C", "no_bc"),
    ("w dt", "w_dt"),
    ("linear interpolation", "linear_interp"),
    ("Geodesic Smoothness Reg", "geo_smooth_reg"),
]

VARIANT_DIRS = {
    "02_baseline_v3": "",
    "V4_tanh_alpha": "tanh_alpha",
    "V4_no_bc": "no_bc",
    "V4_w_dt": "w_dt",
    "V4_linear_interp": "linear_interp",
    "V4_geo_smooth_reg": "geo_smooth_reg",
}

DATASETS = ["ETTh2", "ETTm1", "Weather", "ECL", "PEMS08"]

STANDARD_PRED_LENS = [96, 192, 336, 720]
PEMS_PRED_LENS = [12, 24, 48, 96]

DS_PRED_LENS = {
    "ETTm1": STANDARD_PRED_LENS,
    "Weather": STANDARD_PRED_LENS,
    "ECL": STANDARD_PRED_LENS,
    "ETTh2": STANDARD_PRED_LENS,
    "PEMS08": PEMS_PRED_LENS,
}

DS_TO_COL = {"ETTh2": 3, "ETTm1": 5, "Weather": 7, "ECL": 9, "PEMS08": 11}

_DES_TO_VARIANT = {}
for _name, _vid in VARIANTS:
    if _vid:
        _DES_TO_VARIANT[f"ablation_{_vid}"] = _vid
    else:
        _DES_TO_VARIANT["ablation_V3"] = ""

BASELINE_DES = "ablation_V3"

HEADER_ROWS = 2
DATA_START = 3
N_PL = 4
ROWS_PER_VARIANT = 5
N_VARIANTS = len(VARIANTS)
BLANK_ROW = DATA_START + N_VARIANTS * ROWS_PER_VARIANT
SUMMARY_START = BLANK_ROW + 1

_RE_MSE = re.compile(r"mse:([0-9.eE+-]+)")
_RE_MAE = re.compile(r"mae:([0-9.eE+-]+)")
_RE_PL = re.compile(r"\|pl:(\d+)\|")
_RE_MODEL_ID_PL = re.compile(r"_(\d+)_(\d+)_")


def parse_result_file(filepath):
    results = []
    try:
        with open(filepath, "r") as f:
            lines = f.readlines()
    except FileNotFoundError:
        return results

    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue

        setting = line
        parts = setting.split("|")
        des = parts[1] if len(parts) >= 2 else ""
        pl_match = _RE_PL.search(setting)
        if pl_match:
            pred_len = int(pl_match.group(1))
        else:
            model_id_part = parts[0] if parts else ""
            mid_match = _RE_MODEL_ID_PL.search(model_id_part)
            pred_len = int(mid_match.group(2)) if mid_match else None

        i += 1
        if i >= len(lines):
            break

        metrics_line = lines[i].strip()
        mse_match = _RE_MSE.search(metrics_line)
        mae_match = _RE_MAE.search(metrics_line)
        mse_val = float(mse_match.group(1)) if mse_match else "-"
        mae_val = float(mae_match.group(1)) if mae_match else "-"
        results.append({
            "des": des,
            "pred_len": pred_len,
            "mse": mse_val,
            "mae": mae_val,
        })
        i += 1

    return results


def variant_row(variant_id):
    for i, (_, vid) in enumerate(VARIANTS):
        if vid == variant_id:
            return DATA_START + i * ROWS_PER_VARIANT
    return None


def find_result_files(output_dir):
    file_map = {}
    if not os.path.isdir(output_dir):
        return file_map

    for dir_name, variant_id in VARIANT_DIRS.items():
        dir_path = os.path.join(output_dir, dir_name)
        if not os.path.isdir(dir_path):
            continue

        if variant_id == "":
            des_key = BASELINE_DES
        else:
            des_key = f"ablation_{variant_id}"

        for txt in os.listdir(dir_path):
            if not txt.endswith(".txt"):
                continue
            ds_name = _normalize_ds(txt.replace(".txt", ""))
            if ds_name not in DATASETS:
                continue
            for r in parse_result_file(os.path.join(dir_path, txt)):
                key = (des_key, ds_name, r["pred_len"])
                file_map[key] = r

    return file_map


def _normalize_ds(name):
    name_lower = name.lower().replace("-", "").replace("_", "")
    for ds in DATASETS:
        ds_lower = ds.lower()
        if ds_lower == name_lower:
            return ds
    for ds in DATASETS:
        ds_lower = ds.lower()
        if ds_lower in name_lower or name_lower in ds_lower:
            return ds
    return name


def _unique_sorted_vals(pairs):
    vals = sorted(set(v for v, _ in pairs))
    return vals


def _fmt3(v):
    return math.floor(v * 1000 + 0.5) / 1000


def _is_numeric(v):
    return v is not None and v != "-" and isinstance(v, (int, float))


def _update_avg(ws):
    for vi in range(N_VARIANTS):
        base_row = DATA_START + vi * ROWS_PER_VARIANT
        avg_row = base_row + N_PL

        for ds, col_base in DS_TO_COL.items():
            mse_vals = []
            mae_vals = []
            for pi in range(N_PL):
                row = base_row + pi
                mse = ws.cell(row=row, column=col_base).value
                mae = ws.cell(row=row, column=col_base + 1).value
                if _is_numeric(mse):
                    mse_vals.append(mse)
                if _is_numeric(mae):
                    mae_vals.append(mae)

            if mse_vals:
                c = ws.cell(row=avg_row, column=col_base,
                            value=_fmt3(sum(mse_vals) / len(mse_vals)))
                c.number_format = "0.000"
            if mae_vals:
                c = ws.cell(row=avg_row, column=col_base + 1,
                            value=_fmt3(sum(mae_vals) / len(mae_vals)))
                c.number_format = "0.000"


def _update_summary(ws):
    for pi in range(ROWS_PER_VARIANT):
        for ds, col_base in DS_TO_COL.items():
            pairs_mse = []
            pairs_mae = []
            for vi in range(N_VARIANTS):
                row = DATA_START + vi * ROWS_PER_VARIANT + pi
                mse = ws.cell(row=row, column=col_base).value
                mae = ws.cell(row=row, column=col_base + 1).value
                if _is_numeric(mse):
                    pairs_mse.append((mse, row))
                if _is_numeric(mae):
                    pairs_mae.append((mae, row))

            unique_mse = _unique_sorted_vals(pairs_mse)
            unique_mae = _unique_sorted_vals(pairs_mae)

            summary_row = SUMMARY_START + pi
            if len(unique_mse) >= 1:
                c = ws.cell(row=summary_row, column=col_base, value=unique_mse[0])
                c.number_format = "0.000"
            if len(unique_mae) >= 1:
                c = ws.cell(row=summary_row, column=col_base + 1, value=unique_mae[0])
                c.number_format = "0.000"


def _highlight_best(ws):
    best_fill = PatternFill("solid", fgColor="FFCCCC")
    best_font = Font(name="Arial", bold=True, size=11, color="CC0000")
    default_font = Font(name="Arial", size=11)
    default_fill = PatternFill(fill_type=None)

    max_col = max(DS_TO_COL.values()) + 2

    for vi in range(N_VARIANTS):
        for pi in range(ROWS_PER_VARIANT):
            row = DATA_START + vi * ROWS_PER_VARIANT + pi
            for col in range(3, max_col):
                ws.cell(row=row, column=col).font = default_font
                ws.cell(row=row, column=col).fill = default_fill

    for pi in range(ROWS_PER_VARIANT):
        for ds, col_base in DS_TO_COL.items():
            for col in (col_base, col_base + 1):
                pairs = []
                for vi in range(N_VARIANTS):
                    row = DATA_START + vi * ROWS_PER_VARIANT + pi
                    val = ws.cell(row=row, column=col).value
                    if _is_numeric(val):
                        pairs.append((val, row))
                pairs.sort(key=lambda x: x[0])

                unique_vals = _unique_sorted_vals(pairs)

                if len(unique_vals) >= 1:
                    best_val = unique_vals[0]
                    for v, r in pairs:
                        if v == best_val:
                            ws.cell(row=r, column=col).fill = best_fill
                            ws.cell(row=r, column=col).font = best_font


def update_xlsx(xlsx_path, result_map):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    written = 0
    for (des, ds, pl), metrics in result_map.items():
        vid = _DES_TO_VARIANT.get(des)
        if vid is None:
            continue
        base_row = variant_row(vid)
        if base_row is None:
            continue

        ds_col = DS_TO_COL.get(ds)
        if ds_col is None:
            continue

        pred_lens = DS_PRED_LENS.get(ds, [])
        if pl not in pred_lens:
            continue
        pl_offset = pred_lens.index(pl)

        row = base_row + pl_offset
        mse_val = metrics["mse"]
        mae_val = metrics["mae"]
        c_mse = ws.cell(row=row, column=ds_col,
                        value=_fmt3(mse_val) if mse_val != "-" else "-")
        c_mae = ws.cell(row=row, column=ds_col + 1,
                        value=_fmt3(mae_val) if mae_val != "-" else "-")
        if _is_numeric(c_mse.value):
            c_mse.number_format = "0.000"
        if _is_numeric(c_mae.value):
            c_mae.number_format = "0.000"
        written += 1

    _update_avg(ws)
    _update_summary(ws)
    _highlight_best(ws)
    wb.save(xlsx_path)
    print(f"Updated {written} cells in {xlsx_path}")
    return written


def create_xlsx(xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ManiMamba Ablation v4"

    header_font = Font(name="Arial", bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 8
    for col_idx in range(3, 13):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ds_col_pairs = [
        (3, "ETTh2"),
        (5, "ETTm1"),
        (7, "Weather"),
        (9, "ECL"),
        (11, "PEMS08"),
    ]
    for col_idx, ds_name in ds_col_pairs:
        ws.merge_cells(
            start_row=1, start_column=col_idx,
            end_row=1, end_column=col_idx + 1,
        )
        cell = ws.cell(row=1, column=col_idx, value=ds_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.cell(row=1, column=col_idx + 1).border = thin_border

    sub_headers = {}
    for col_idx, _ in ds_col_pairs:
        sub_headers[col_idx] = "MSE"
        sub_headers[col_idx + 1] = "MAE"
    for col_idx, label in sub_headers.items():
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = center_align
        cell.border = thin_border

    ws.cell(row=1, column=1, value="Variant").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border
    ws.cell(row=2, column=1).border = thin_border
    ws.cell(row=1, column=2, value="Len").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = center_align
    ws.cell(row=1, column=2).border = thin_border
    ws.cell(row=2, column=2).border = thin_border

    for vi, (vname, _vid) in enumerate(VARIANTS):
        base_row = DATA_START + vi * ROWS_PER_VARIANT
        ws.merge_cells(
            start_row=base_row, start_column=1,
            end_row=base_row + N_PL, end_column=1,
        )
        cell = ws.cell(row=base_row, column=1, value=vname)
        cell.font = Font(name="Arial", bold=True, size=11)
        cell.alignment = Alignment(vertical="center")
        for pi in range(ROWS_PER_VARIANT):
            row = base_row + pi
            if pi < N_PL:
                ws.cell(row=row, column=2,
                        value=f"{STANDARD_PRED_LENS[pi]}/{PEMS_PRED_LENS[pi]}")
            else:
                ws.cell(row=row, column=2, value="Avg")
            ws.cell(row=row, column=2).alignment = center_align
            for col_idx in range(1, 13):
                ws.cell(row=row, column=col_idx).border = thin_border

    for pi in range(ROWS_PER_VARIANT):
        row = SUMMARY_START + pi
        if pi == 0:
            ws.cell(row=row, column=1, value="Min")
            ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=11)
        if pi < N_PL:
            ws.cell(row=row, column=2,
                    value=f"{STANDARD_PRED_LENS[pi]}/{PEMS_PRED_LENS[pi]}")
        else:
            ws.cell(row=row, column=2, value="Avg")
        ws.cell(row=row, column=2).alignment = center_align
        for col_idx in range(1, 13):
            ws.cell(row=row, column=col_idx).border = thin_border

    ws.freeze_panes = "C3"

    wb.save(xlsx_path)
    print(f"Created {xlsx_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Update ManiMamba v4 ablation xlsx from experiment results")
    parser.add_argument(
        "--xlsx", default="output/ablation/manimamba-ablation-v4.xlsx",
        help="Path to xlsx file")
    parser.add_argument(
        "--output-dir", default="output/ablation",
        help="Root output directory for ablation results")
    parser.add_argument(
        "--create", action="store_true",
        help="Create xlsx from scratch instead of updating")
    args = parser.parse_args()

    if args.create:
        create_xlsx(args.xlsx)
        return

    if not os.path.isfile(args.xlsx):
        print(f"xlsx not found, creating: {args.xlsx}")
        create_xlsx(args.xlsx)

    result_map = find_result_files(args.output_dir)
    if not result_map:
        print(f"No results found under {args.output_dir}")
        print("Expected structure: output/ablation/"
              "{02_baseline_v3,V4_tanh_alpha,...}/{DATASET}.txt")
        sys.exit(1)

    print(f"Found {len(result_map)} result entries")
    update_xlsx(args.xlsx, result_map)


if __name__ == "__main__":
    main()
