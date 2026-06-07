#!/usr/bin/env python3
"""
Read Optuna best_params JSON files from output/optuna/ and update all
ManiMamba experiment Excel files matching assets/mani_eff_exp_*.xlsx with:
    - MSE / MAE in the Experiment Results section (columns C-D)
    - Best hyperparameters in the Config section (columns C-W)

Usage:
    python scripts/update_excel_from_optuna.py [--dry-run]
"""

import argparse
import json
import re
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ASSETS_DIR = PROJECT_ROOT / "assets"
OPTUNA_DIR = PROJECT_ROOT / "output" / "optuna"
ARCHIVE_DIR = OPTUNA_DIR / "archive"

METRIC_NUM_FMT = "0.000"

SHEET_NAME = "MODE_Experiment_Corrected_Formu"

RESULTS_HEADER_ROW = 3
CONFIG_HEADER_ROW = 67
RESULTS_DATA_START = 4
RESULTS_DATA_END = 63
CONFIG_DATA_START = 68
CONFIG_DATA_END = 115

PARAM_COL_MAP = {
    "e_layers": "D",
    "batch_size": "E",
    "batch": "E",
    "d_model": "F",
    "d_state": "G",
    "expand": "H",
    "dropout": "I",
    "epsilon": "J",
    "cov_window": "K",
    "cov_stride": "L",
    "cov_rank": "M",
    "geo_d_model": "N",
    "geo_d_state": "O",
    "geo_d_conv": "P",
    "geo_expand": "Q",
    "d_ff": "R",
    "warmup": "S",
    "warmup_epochs": "S",
    "epochs": "T",
    "train_epochs": "T",
    "patience": "U",
    "weight_decay": "V",
    "learning_rate": "W",
}

DATASET_ALIASES = {
    "Solar": "Solar-Energy",
    "Electricity": "ECL",
    "illness": "Illness",
}

DATASET_PRED_LENS = {
    "ETTm1": [96, 192, 336, 720],
    "ETTm2": [96, 192, 336, 720],
    "ETTh1": [96, 192, 336, 720],
    "ETTh2": [96, 192, 336, 720],
    "Weather": [96, 192, 336, 720],
    "ECL": [96, 192, 336, 720],
    "Traffic": [96, 192, 336, 720],
    "Exchange": [96, 192, 336, 720],
    "Solar-Energy": [96, 192, 336, 720],
    "PEMS03": [12, 24, 48, 96],
    "PEMS04": [12, 24, 48, 96],
    "PEMS07": [12, 24, 48, 96],
    "PEMS08": [12, 24, 48, 96],
    "Illness": [24, 36, 48, 60],
}


def round_metric(value):
    return round(float(value), 3)


def col_to_idx(col: str) -> int:
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def build_row_map(ws, start_row, end_row):
    """Scan columns A-B to build {(dataset, pred_len): row_number}."""
    row_map = {}
    current_dataset = None
    for row in range(start_row, end_row + 1):
        a_val = ws.cell(row=row, column=1).value
        b_val = ws.cell(row=row, column=2).value
        if a_val is not None:
            current_dataset = a_val
        if current_dataset and b_val is not None and isinstance(b_val, (int, float)):
            key = (current_dataset, int(b_val))
            row_map[key] = row
    return row_map


def parse_json_filename(fname: str):
    stem = Path(fname).stem
    m = re.match(r"ManiMamba_(.+)_pl(\d+)_best_params", stem)
    if not m:
        return None, None
    dataset = m.group(1)
    pred_len = int(m.group(2))
    return dataset, pred_len


def find_json_files(optuna_dir: Path) -> list[Path]:
    files = []
    if not optuna_dir.exists():
        return files
    for f in sorted(optuna_dir.glob("ManiMamba_*_best_params.json")):
        if ARCHIVE_DIR in f.parents:
            continue
        files.append(f)
    return files


def resolve_dataset_name(raw_name: str) -> str:
    return DATASET_ALIASES.get(raw_name, raw_name)


def find_excel_files(assets_dir: Path) -> list[Path]:
    return sorted(assets_dir.glob("mani_eff_exp_*.xlsx"))


def fill_averages(ws, dry_run: bool = False) -> int:
    """Scan Avg rows; if all pred_len rows for a dataset have MSE+MAE, write the average."""
    filled = 0
    for row in range(RESULTS_DATA_START, RESULTS_DATA_END + 1):
        if ws.cell(row=row, column=2).value != "Avg":
            continue
        dataset = None
        for r in range(row - 1, RESULTS_DATA_START - 1, -1):
            a = ws.cell(row=r, column=1).value
            if a is not None:
                dataset = a
                break
        if not dataset:
            continue
        num_rows = len(DATASET_PRED_LENS.get(dataset, [96, 192, 336, 720]))
        data_start = row - num_rows

        for col in range(3, 21):
            metric_label = ws.cell(row=RESULTS_HEADER_ROW, column=col).value
            if metric_label not in ("MSE", "MAE"):
                continue
            vals = []
            for dr in range(data_start, row):
                v = ws.cell(row=dr, column=col).value
                if v is not None:
                    vals.append(float(v))
            if len(vals) != num_rows:
                continue
            avg = round_metric(sum(vals) / len(vals))
            old = ws.cell(row=row, column=col).value
            if not dry_run:
                c = ws.cell(row=row, column=col)
                c.value = avg
                c.number_format = METRIC_NUM_FMT
            filled += 1
        print(f"  Avg row {row} ({dataset}): C={ws.cell(row=row, column=3).value}")
    return filled


def update_single_excel(
    excel_path: Path, json_files: list[Path], dry_run: bool = False
) -> int:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[SHEET_NAME]

    results_map = build_row_map(ws, RESULTS_DATA_START, RESULTS_DATA_END)
    config_map = build_row_map(ws, CONFIG_DATA_START, CONFIG_DATA_END)

    updated = 0
    for jf in json_files:
        dataset_raw, pred_len = parse_json_filename(jf.name)
        if dataset_raw is None:
            print(f"  SKIP (bad filename): {jf.name}")
            continue

        dataset = resolve_dataset_name(dataset_raw)

        with open(jf) as f:
            data = json.load(f)

        mse = data.get("best_trial_user_attrs", {}).get("mse") or data.get("best_value")
        mae = data.get("best_trial_user_attrs", {}).get("mae")
        params = data.get("best_params", {})
        version = data.get("version", "")

        if mse is None or mae is None:
            print(f"  SKIP (missing mse/mae): {jf.name}")
            continue

        version_label = version if version else "legacy"

        results_key = (dataset, pred_len)
        config_key = (dataset, pred_len)

        print(f"\n{jf.name}  [version={version_label}]")
        print(f"  Dataset={dataset}  pred_len={pred_len}  MSE={round_metric(mse):.3f}  MAE={round_metric(mae):.3f}")

        if results_key in results_map:
            row = results_map[results_key]
            old_mse = ws.cell(row=row, column=col_to_idx("C")).value
            old_mae = ws.cell(row=row, column=col_to_idx("D")).value
            if not dry_run:
                c = ws.cell(row=row, column=col_to_idx("C"))
                c.value = round_metric(mse)
                c.number_format = METRIC_NUM_FMT
                d = ws.cell(row=row, column=col_to_idx("D"))
                d.value = round_metric(mae)
                d.number_format = METRIC_NUM_FMT
            action = "OVERWRITE" if old_mse is not None else "NEW"
            print(
                f"  Results row {row}: [{action}] MSE {old_mse} -> {round_metric(mse):.3f}, MAE {old_mae} -> {round_metric(mae):.3f}"
            )
        else:
            print(f"  WARNING: no matching results row for ({dataset}, {pred_len})")

        if config_key in config_map:
            row = config_map[config_key]
            param_str_parts = []
            for param_name, col_letter in PARAM_COL_MAP.items():
                if param_name in params:
                    val = params[param_name]
                    col_idx = col_to_idx(col_letter)
                    if not dry_run:
                        ws.cell(row=row, column=col_idx).value = val
                    param_str_parts.append(f"{param_name}={val}")
            print(f"  Config row {row}: updated {len(param_str_parts)} params")
        else:
            print(f"  WARNING: no matching config row for ({dataset}, {pred_len})")

        updated += 1

    if not dry_run and (updated > 0):
        print(f"\n--- Filling averages ---")
        avg_filled = fill_averages(ws, dry_run=dry_run)
        print(f"  Avg rows filled: {avg_filled}")

    if updated > 0 and not dry_run:
        wb.save(excel_path)
        print(f"\nExcel saved: {excel_path}")
        print(f"Total entries updated: {updated}")
    elif dry_run and updated > 0:
        print(f"\n[DRY RUN] Would update {updated} entries. No changes saved.")
    else:
        print("  No matching entries to update.")

    return updated


def update_excel(dry_run: bool = False):
    excel_files = find_excel_files(ASSETS_DIR)
    if not excel_files:
        print(
            f"ERROR: No Excel files matching 'mani_eff_exp_*.xlsx' found in {ASSETS_DIR}"
        )
        return

    json_files = find_json_files(OPTUNA_DIR)
    if not json_files:
        print("No Optuna JSON files found (excluding archive).")
        return

    total_updated = 0
    for excel_path in excel_files:
        print(f"\n{'='*60}")
        print(f"Processing: {excel_path.name}")
        print(f"{'='*60}")
        total_updated += update_single_excel(excel_path, json_files, dry_run=dry_run)

    print(f"\nAll Excel files processed. Total entries updated: {total_updated}")


def main():
    parser = argparse.ArgumentParser(
        description="Update experiment Excel from Optuna JSON results"
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Preview changes without saving"
    )
    args = parser.parse_args()
    update_excel(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
