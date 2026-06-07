#!/usr/bin/env python3
"""
Unified efficiency experiment runner for ManiMamba model.

Usage:
    python scripts/efficiency/run_efficiency.py --dataset ECL
    python scripts/efficiency/run_efficiency.py --dataset ETTm1 --pred_len 96
    python scripts/efficiency/run_efficiency.py --dataset PEMS04 --expand 1
    python scripts/efficiency/run_efficiency.py --dataset Weather --train_epochs 10

Any unrecognized args are passed through to scripts/run.py.
Results are written to:
  - output/efficiency/manimamba_results/ManiMamba_{Dataset}_pl{pred_len}_{timestamp}.json
  - output/efficiency/manimamba_results/efficiency_records.xlsx (updated or created)
"""

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import yaml

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
YAML_PATH = os.path.join(SCRIPT_DIR, "efficiency_defaults.yaml")
DEFAULT_OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output", "efficiency", "manimamba_results")
XLSX_PATH = os.path.join(DEFAULT_OUTPUT_DIR, "efficiency_records.xlsx")

XLSX_HEADERS = [
    "Timestamp", "Dataset", "Model", "Model ID",
    "seq_len", "pred_len", "enc_in", "e_layers",
    "d_model", "d_ff", "d_state", "expand",
    "epsilon", "cov_window", "cov_stride", "cov_rank",
    "geo_d_model", "geo_d_state", "geo_d_conv", "geo_expand",
    "geo_inject_threshold", "ablation",
    "dropout",
    "learning_rate", "weight_decay", "batch_size", "optim",
    "train_epochs", "warmup_epochs", "patience", "use_amp", "gpu",
    "MSE", "MAE", "Train Time (ms/iter)", "GPU Memory (MB)",
]

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def load_defaults(dataset, pred_len=None):
    """Load ManiMamba default config from YAML for the given dataset and pred_len.

    Lookup is case-insensitive on dataset name.  The YAML key format is
    ``{Dataset}_{pred_len}`` (e.g. ``ECL_336``, ``Weather_96``).
    """
    with open(YAML_PATH) as f:
        all_defaults = yaml.safe_load(f)

    mani_mamba = all_defaults.get("ManiMamba", {})
    dataset_lower = dataset.lower()

    index = {}
    for key, config in mani_mamba.items():
        parts = key.rsplit("_", 1)
        if len(parts) != 2:
            continue
        ds, pl = parts[0], int(parts[1])
        ds_lower = ds.lower()
        if ds_lower not in index:
            index[ds_lower] = {}
        index[ds_lower][pl] = config

    ds_configs = index.get(dataset_lower, {})

    if pred_len is not None:
        cfg = ds_configs.get(pred_len)
        if cfg is None:
            available = list(ds_configs.keys())
            print("ERROR: No config for {} pred_len={}".format(dataset, pred_len))
            print("  Available pred_lens: {}".format(available))
            sys.exit(1)
        return dict(cfg)

    if len(ds_configs) == 0:
        print("ERROR: No config found for dataset '{}'".format(dataset))
        print("  Available datasets: {}".format(list(index.keys())))
        sys.exit(1)

    if len(ds_configs) == 1:
        return dict(list(ds_configs.values())[0])

    print("Multiple configs for {}: pred_lens {}".format(dataset, list(ds_configs.keys())))
    print("Specify --pred_len to disambiguate.")
    sys.exit(1)


def apply_overrides(config, args):
    """Override YAML config values with non-None CLI arguments."""
    override_keys = [
        "seq_len", "pred_len",
        "e_layers", "d_model", "d_ff",
        "d_state", "expand",
        "epsilon", "cov_window", "cov_stride", "cov_rank",
        "geo_d_model", "geo_d_state", "geo_d_conv", "geo_expand",
        "geo_inject_threshold",
        "ablation",
        "batch_size", "learning_rate", "weight_decay", "dropout",
        "patience", "warmup_epochs", "lradj", "cosine_eta_min", "optim", "use_amp",
        "gpu",
    ]
    for key in override_keys:
        val = getattr(args, key, None)
        if val is not None:
            config[key] = val
    if args.type is not None:
        config["_result_type"] = args.type
    return config


def build_run_command(config, train_epochs, extra_args):
    """Build the subprocess command list for ``scripts/run.py``."""
    model_id = "{}_{}_{}".format(config["dataset"], config["seq_len"], config["pred_len"])

    cmd = [
        sys.executable, "-u",
        os.path.join(PROJECT_ROOT, "scripts", "run.py"),
        "--is_training", "1",
        "--model", "ManiMamba",
        "--model_id", model_id,
        "--root_path", config["root_path"],
        "--data_path", config["data_path"],
        "--data", config["data"],
        "--freq", config.get("freq", "h"),
        "--features", "M",
        "--seq_len", str(config["seq_len"]),
        "--pred_len", str(config["pred_len"]),
        "--enc_in", str(config["enc_in"]),
        "--dec_in", str(config["dec_in"]),
        "--c_out", str(config["c_out"]),
        "--e_layers", str(config["e_layers"]),
        "--d_model", str(config["d_model"]),
        "--d_ff", str(config["d_ff"]),
        "--des", "Efficiency-Exp",
        "--itr", "1",
        "--batch_size", str(config["batch_size"]),
        "--train_epochs", str(train_epochs),
        "--learning_rate", str(config["learning_rate"]),
        "--dropout", str(config.get("dropout", 0.2)),
        "--patience", str(config.get("patience", 7)),
        "--use_amp" if config.get("use_amp", 1) else "--no-use_amp",
        "--optim", config.get("optim", "AdamW"),
        "--weight_decay", str(config.get("weight_decay", 1e-5)),
        "--lradj", config.get("lradj", "type1"),
        "--cosine_eta_min", str(config.get("cosine_eta_min", 1e-7)),
        "--warmup_epochs", str(config.get("warmup_epochs", 5)),
        "--gpu", str(config.get("gpu", 0)),
        "--d_state", str(config["d_state"]),
        "--expand", str(config["expand"]),
        "--epsilon", str(config["epsilon"]),
        "--cov_window", str(config["cov_window"]),
        "--cov_stride", str(config["cov_stride"]),
        "--cov_rank", str(config["cov_rank"]),
        "--geo_d_model", str(config["geo_d_model"]),
        "--geo_d_state", str(config["geo_d_state"]),
        "--geo_d_conv", str(config["geo_d_conv"]),
        "--geo_expand", str(config["geo_expand"]),
        "--geo_inject_threshold", str(config.get("geo_inject_threshold", 100)),
        "--use_cuda_accel", str(config.get("use_cuda_accel", 0)),
    ]

    abl = config.get("ablation", "")
    if abl:
        cmd.extend(["--ablation", abl])

    cmd.extend(extra_args)
    return cmd, model_id


def parse_training_output(log_content):
    """Parse training log and extract MSE, MAE, training speed, GPU memory."""
    metrics = {
        "mse": 0.0,
        "mae": 0.0,
        "train_time_ms_per_iter": 0.0,
        "avg_gpu_mem_allocated_mb": 0.0,
    }

    m = re.search(r"mse:([0-9\.e\-]+),\s*mae:([0-9\.e\-]+)", log_content)
    if m:
        metrics["mse"] = float(m.group(1))
        metrics["mae"] = float(m.group(2))

    speeds = re.findall(r"speed: ([0-9\.]+)s/iter", log_content)
    if speeds:
        times_ms = [float(t) * 1000 for t in speeds]
        metrics["train_time_ms_per_iter"] = sum(times_ms) / len(times_ms)

    gpu_matches = re.findall(r"allocated_memory: ([0-9\.]+)", log_content)
    if gpu_matches:
        vals = [float(g) * 1024 for g in gpu_matches]
        metrics["avg_gpu_mem_allocated_mb"] = sum(vals) / len(vals)

    return metrics


def build_result_json(model_id, timestamp, config, metrics):
    """Build the JSON-serializable result dict."""
    ds = config.get("dataset", "unknown")
    result = {}
    if config.get("_result_type"):
        result["type"] = config["_result_type"]
    result["model"] = "ManiMamba"
    result["dataset"] = ds
    result["model_id"] = model_id
    result["timestamp"] = timestamp
    result["metrics"] = {
        "mse": metrics["mse"],
        "mae": metrics["mae"],
        "train_time_ms_per_iter": round(metrics["train_time_ms_per_iter"], 4),
        "avg_gpu_mem_allocated_mb": round(metrics["avg_gpu_mem_allocated_mb"], 4),
    }
    result["config"] = {
        "seq_len": config.get("seq_len"),
        "pred_len": config.get("pred_len"),
        "enc_in": config.get("enc_in"),
        "e_layers": config.get("e_layers"),
        "d_model": config.get("d_model"),
        "d_ff": config.get("d_ff"),
        "d_state": config.get("d_state"),
        "expand": config.get("expand"),
        "epsilon": config.get("epsilon"),
        "cov_window": config.get("cov_window"),
        "cov_stride": config.get("cov_stride"),
        "cov_rank": config.get("cov_rank"),
        "geo_d_model": config.get("geo_d_model"),
        "geo_d_state": config.get("geo_d_state"),
        "geo_d_conv": config.get("geo_d_conv"),
        "geo_expand": config.get("geo_expand"),
        "geo_inject_threshold": config.get("geo_inject_threshold"),
        "ablation": config.get("ablation", ""),
        "dropout": config.get("dropout"),
        "learning_rate": config.get("learning_rate"),
        "weight_decay": config.get("weight_decay"),
        "batch_size": config.get("batch_size"),
        "optim": config.get("optim"),
        "train_epochs": config.get("_train_epochs"),
        "warmup_epochs": config.get("warmup_epochs"),
        "patience": config.get("patience"),
        "use_amp": config.get("use_amp"),
        "gpu": config.get("gpu", 0),
    }
    return result


def write_json(output_dir, result):
    """Write result as a JSON file to output_dir."""
    ds = result["dataset"]
    pl = result["config"]["pred_len"]
    ts = result["timestamp"]
    json_path = os.path.join(
        output_dir, "ManiMamba_{}_pl{}_{}.json".format(ds, pl, ts))
    with open(json_path, "w") as f:
        json.dump(result, f, indent=4)
    return json_path


def _row_values_from_result(result):
    """Extract ordered row values matching XLSX_HEADERS."""
    c = result["config"]
    m = result["metrics"]
    return [
        result["timestamp"],
        result["dataset"],
        result["model"],
        result["model_id"],
        c.get("seq_len"),
        c.get("pred_len"),
        c.get("enc_in"),
        c.get("e_layers"),
        c.get("d_model"),
        c.get("d_ff"),
        c.get("d_state"),
        c.get("expand"),
        c.get("epsilon"),
        c.get("cov_window"),
        c.get("cov_stride"),
        c.get("cov_rank"),
        c.get("geo_d_model"),
        c.get("geo_d_state"),
        c.get("geo_d_conv"),
        c.get("geo_expand"),
        c.get("geo_inject_threshold"),
        c.get("ablation"),
        c.get("dropout"),
        c.get("learning_rate"),
        c.get("weight_decay"),
        c.get("batch_size"),
        c.get("optim"),
        c.get("train_epochs"),
        c.get("warmup_epochs"),
        c.get("patience"),
        c.get("use_amp"),
        c.get("gpu"),
        m["mse"],
        m["mae"],
        m["train_time_ms_per_iter"],
        m["avg_gpu_mem_allocated_mb"],
    ]


def update_xlsx(result):
    """Append a result row to the efficiency_records.xlsx tracking file.

    Creates the file with headers if it does not exist.
    """
    if os.path.exists(XLSX_PATH):
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ManiMamba Efficiency"
        for col_idx, header in enumerate(XLSX_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        start_row = 2

    values = _row_values_from_result(result)
    for col_idx, val in enumerate(values, 1):
        ws.cell(row=start_row, column=col_idx, value=val)

    ws.auto_filter.ref = ws.dimensions
    wb.save(XLSX_PATH)
    wb.close()


def main():
    parser = argparse.ArgumentParser(
        description="Run ManiMamba efficiency experiments",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Any additional unrecognized args are passed through to scripts/run.py.\n"
            "Examples:\n"
            "  python scripts/efficiency/run_efficiency.py --dataset ECL\n"
            "  python scripts/efficiency/run_efficiency.py --dataset ETTm1 --pred_len 96\n"
            "  python scripts/efficiency/run_efficiency.py --dataset PEMS04 --expand 1\n"
            "  python scripts/efficiency/run_efficiency.py --dataset Weather --train_epochs 10\n"
        ),
    )

    parser.add_argument("--dataset", type=str, required=True,
                        help="Dataset name: ECL, ETTm1, PEMS04, Weather")
    parser.add_argument("--pred_len", type=int, default=None,
                        help="Prediction length (default: from YAML config)")
    parser.add_argument("--train_epochs", type=int, default=10,
                        help="Number of training epochs")
    parser.add_argument("--gpu", type=int, default=None, help="GPU device id")
    parser.add_argument("--output_dir", type=str, default=None,
                        help="Output directory for results (default: output/efficiency/manimamba_results/)")

    parser.add_argument("--seq_len", type=int, default=None)
    parser.add_argument("--e_layers", type=int, default=None)
    parser.add_argument("--d_model", type=int, default=None)
    parser.add_argument("--d_ff", type=int, default=None)
    parser.add_argument("--batch_size", type=int, default=None)
    parser.add_argument("--learning_rate", type=float, default=None)
    parser.add_argument("--weight_decay", type=float, default=None)
    parser.add_argument("--dropout", type=float, default=None)
    parser.add_argument("--patience", type=int, default=None)
    parser.add_argument("--warmup_epochs", type=int, default=None)
    parser.add_argument("--lradj", type=str, default=None)
    parser.add_argument("--cosine_eta_min", type=float, default=None)
    parser.add_argument("--optim", type=str, default=None)
    parser.add_argument("--use_amp", type=int, default=None)

    parser.add_argument("--d_state", type=int, default=None)
    parser.add_argument("--expand", type=int, default=None)
    parser.add_argument("--epsilon", type=float, default=None)
    parser.add_argument("--cov_window", type=int, default=None)
    parser.add_argument("--cov_stride", type=int, default=None)
    parser.add_argument("--cov_rank", type=int, default=None)
    parser.add_argument("--geo_d_model", type=int, default=None)
    parser.add_argument("--geo_d_state", type=int, default=None)
    parser.add_argument("--geo_d_conv", type=int, default=None)
    parser.add_argument("--geo_expand", type=int, default=None)
    parser.add_argument("--geo_inject_threshold", type=int, default=None)
    parser.add_argument("--ablation", type=str, default=None,
                        choices=["tanh_alpha", "no_bc", "w_dt", "linear_interp", "geo_smooth_reg"])
    parser.add_argument("--type", type=str, default=None,
                        help="Result type label (e.g. 'comparison'). Only written to JSON when set.")

    args, extra_run_args = parser.parse_known_args()

    config = load_defaults(args.dataset, args.pred_len)
    config = apply_overrides(config, args)
    config["_train_epochs"] = args.train_epochs

    output_dir = args.output_dir or DEFAULT_OUTPUT_DIR
    cmd, model_id = build_run_command(config, args.train_epochs, extra_run_args)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("Config: batch={}, epochs={}, lr={}, dm={}, df={}, el={}, dstate={}, exp={}, eps={}, cw={}, cs={}, cr={}, gdm={}, gds={}, gdc={}, gexp={}".format(
        config["batch_size"], args.train_epochs, config["learning_rate"],
        config["d_model"], config["d_ff"], config["e_layers"],
        config["d_state"], config["expand"],
        config["epsilon"],
        config["cov_window"], config["cov_stride"], config["cov_rank"],
        config["geo_d_model"], config["geo_d_state"],
        config["geo_d_conv"], config["geo_expand"],
    ))
    print("Model ID: {}".format(model_id))
    print("=" * 80)
    print()
    print("Starting training...")
    print()

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        cwd=PROJECT_ROOT,
    )

    output_lines = []
    for line in iter(proc.stdout.readline, b""):
        decoded = line.decode("utf-8", errors="replace")
        print(decoded, end="")
        output_lines.append(decoded)

    proc.wait()

    if proc.returncode != 0:
        print("\nERROR: Training failed for ManiMamba on {}".format(args.dataset))
        sys.exit(1)

    log_content = "".join(output_lines)
    print("\nParsing metrics...")

    metrics = parse_training_output(log_content)

    dataset_name = config.get("dataset", args.dataset)
    os.makedirs(output_dir, exist_ok=True)

    result = build_result_json(model_id, timestamp, config, metrics)

    json_path = write_json(output_dir, result)
    update_xlsx(result)

    print("Successfully parsed metrics for ManiMamba on {}".format(dataset_name))
    print("  Model ID: {}".format(model_id))
    print("  MSE: {:.8f}".format(metrics["mse"]))
    print("  MAE: {:.8f}".format(metrics["mae"]))
    print("  Training Time: {:.2f} ms/iter".format(metrics["train_time_ms_per_iter"]))
    print("  Avg Allocated GPU Memory: {:.2f} MB".format(metrics["avg_gpu_mem_allocated_mb"]))
    print("JSON saved to: {}".format(json_path))
    print("XLSX updated: {}".format(XLSX_PATH))
    print("=" * 80)


if __name__ == "__main__":
    main()
