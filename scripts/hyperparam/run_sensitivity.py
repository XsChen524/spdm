#!/usr/bin/env python3
"""
Hyperparameter sensitivity analysis for ManiMamba (Weather).

Single-experiment runner: each invocation trains and tests exactly one
(sweep_param, pred_len, sweep_value) configuration.  All non-swept
hyperparameters come from DEFAULT_PARAMS below — edit directly.

Usage:
    python scripts/hyperparam/run_sensitivity.py --sweep_param d_model --sweep_value 128 --pred_len 96 --gpu 0
"""

import argparse
import gc
import os
import random
import sys
import time
import traceback

import numpy as np

sys.path.append(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
)

from configs.model_parser_dict import model_parser_dict

DATASET = "Weather"
ENC_IN = 21
DATA_PATH = "weather.csv"
ROOT_PATH = "./data/weather/"
DATA = "custom"

OUTPUT_DIR = os.path.join("output", "hyperparam")
RESULTS_XLSX = os.path.join(OUTPUT_DIR, "sensitivity_results.xlsx")

SWEEP_RANGES = {
    "d_model": [128, 256, 512, 1024],
    "epsilon": [1e-5, 1e-4, 1e-3, 1e-2],
    "geo_d_model": [64, 128, 256, 512],
    "cov_rank": [0, 8, 16, 32],
}

DEFAULT_PARAMS = {
    96: {
        "e_layers": 3,
        "d_model": 256,
        "d_state": 16,
        "expand": 1,
        "epsilon": 1e-4,
        "cov_window": 32,
        "cov_stride": 16,
        "cov_rank": 0,
        "geo_d_model": 128,
        "geo_d_state": 8,
        "geo_d_conv": 2,
        "geo_expand": 1,
        "learning_rate": 4.89e-5,
        "weight_decay": 1e-6,
        "dropout": 0.2,
        "batch_size": 32,
        "train_epochs": 15,
        "warmup_epochs": 3,
        "patience": 5,
    },
    336: {
        "e_layers": 2,
        "d_model": 256,
        "d_state": 2,
        "expand": 1,
        "epsilon": 1e-3,
        "cov_window": 16,
        "cov_stride": 16,
        "cov_rank": 0,
        "geo_d_model": 128,
        "geo_d_state": 8,
        "geo_d_conv": 2,
        "geo_expand": 1,
        "learning_rate": 7.88e-5,
        "weight_decay": 1e-6,
        "dropout": 0.2,
        "batch_size": 32,
        "train_epochs": 15,
        "warmup_epochs": 3,
        "patience": 5,
    },
}

XLSX_HEADER = [
    "dataset",
    "pred_len",
    "sweep_param",
    "sweep_value",
    "mse",
    "mae",
    "status",
    "elapsed_s",
    "timestamp",
]


def _build_default_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(conflict_handler="resolve")
    for p in model_parser_dict["__all__"]:
        p(parser=parser)
    for p in model_parser_dict["ManiMamba"]:
        p(parser=parser)
    return parser.parse_args(
        [
            "--is_training",
            "1",
            "--model_id",
            "test",
            "--model",
            "ManiMamba",
            "--data",
            "custom",
        ]
    )


def _build_setting(args) -> str:
    return "{}_{}|{}|batch:{}|epochs:{}|expand:{}|eps:{}|cw:{}|cs:{}|cr:{}|gdm:{}|gds:{}|gdc:{}|ge:{}|ds:{}|el:{}|dm:{}|lr:{}|itr:0".format(
        args.model_id,
        args.model,
        args.des,
        args.batch_size,
        args.train_epochs,
        args.expand,
        args.epsilon,
        args.cov_window,
        args.cov_stride,
        args.cov_rank,
        args.geo_d_model,
        args.geo_d_state,
        args.geo_d_conv,
        args.geo_expand,
        args.d_state,
        args.e_layers,
        args.d_model,
        args.learning_rate,
    )


def _write_xlsx_row(row: dict):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    if os.path.exists(RESULTS_XLSX):
        wb = openpyxl.load_workbook(RESULTS_XLSX)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sensitivity_results"
        for col_idx, h in enumerate(XLSX_HEADER, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    next_row = ws.max_row + 1
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col_idx, h in enumerate(XLSX_HEADER, 1):
        val = row[h]
        if h in ("pred_len", "mse", "mae", "elapsed_s", "sweep_value"):
            try:
                val = float(val)
            except (ValueError, TypeError):
                pass
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        if next_row % 2 == 0:
            cell.fill = alt_fill
        if h in ("mse", "mae"):
            cell.number_format = "0.00000000"
        elif h == "elapsed_s":
            cell.number_format = "0.0"
        elif h == "sweep_value":
            sv = val if isinstance(val, (int, float)) else 0
            if isinstance(sv, float) and sv < 0.01:
                cell.number_format = "0.0E+00"
            else:
                cell.number_format = "0.####"

    wb.save(RESULTS_XLSX)


def _apply_params(args: argparse.Namespace, params: dict) -> None:
    for k, v in params.items():
        setattr(args, k, type(v)(v))


def main():
    cli = argparse.ArgumentParser(description="Single-experiment hyperparameter sweep")
    cli.add_argument("--sweep_param", type=str, required=True)
    cli.add_argument("--sweep_value", type=str, required=True)
    cli.add_argument("--pred_len", type=int, required=True)
    cli.add_argument("--gpu", type=int, default=0)
    opts = cli.parse_args()

    param = opts.sweep_param
    if param not in SWEEP_RANGES:
        print(
            f"ERROR: Unknown sweep_param '{param}'. Valid: {list(SWEEP_RANGES.keys())}"
        )
        sys.exit(1)

    if opts.pred_len not in DEFAULT_PARAMS:
        print(
            f"ERROR: No DEFAULT_PARAMS for pred_len={opts.pred_len}. "
            f"Available: {list(DEFAULT_PARAMS.keys())}"
        )
        sys.exit(1)

    raw = opts.sweep_value
    try:
        value = int(raw) if "." not in raw and "e" not in raw.lower() else float(raw)
    except ValueError:
        value = raw

    import torch

    defaults = DEFAULT_PARAMS[opts.pred_len]

    args = _build_default_args()
    args.is_training = 1
    args.model = "ManiMamba"
    args.model_id = f"{DATASET}_96_{opts.pred_len}"
    args.data = DATA
    args.root_path = ROOT_PATH
    args.data_path = DATA_PATH
    args.enc_in = ENC_IN
    args.dec_in = ENC_IN
    args.c_out = ENC_IN
    args.features = "M"
    args.seq_len = 96
    args.pred_len = opts.pred_len
    sv = f"{value:.2e}" if isinstance(value, float) else str(value)
    args.des = f"HPARAM_SENSITIVITY_{param}_{sv}"
    args.use_gpu = True
    args.gpu = opts.gpu
    args.use_multi_gpu = False
    args.devices = str(opts.gpu)
    args.itr = 1
    args.output_attention = False
    args.do_predict = False
    args.save_tmp = 1
    args.exp_name = "MTSF"
    args.use_norm = True

    _apply_params(args, defaults)
    setattr(args, param, type(defaults[param])(value))

    seed = getattr(args, "seed", 2023)
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)

    from src.experiments.exp_long_term_forecasting import Exp_Long_Term_Forecast

    sv = f"{value:.2e}" if isinstance(value, float) else str(value)
    tag = f"[{DATASET} pl={opts.pred_len} {param}={sv}]"
    setting = _build_setting(args)

    row = {
        "dataset": DATASET,
        "pred_len": opts.pred_len,
        "sweep_param": param,
        "sweep_value": value,
        "mse": "",
        "mae": "",
        "status": "running",
        "elapsed_s": "",
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
    }

    print(f"{tag} Starting training ...")
    t0 = time.time()
    exp = None
    try:
        exp = Exp_Long_Term_Forecast(args)
        exp.train(setting)
        test_result = exp.test(setting)

        mse_val = mae_val = None
        if test_result is not None:
            mse_val, mae_val = test_result.get("mse"), test_result.get("mae")
        elif hasattr(exp, "_last_test_result") and exp._last_test_result:
            mse_val = exp._last_test_result.get("mse")
            mae_val = exp._last_test_result.get("mae")

        if (
            mse_val is not None
            and mae_val is not None
            and np.isfinite(mse_val)
            and np.isfinite(mae_val)
        ):
            row["mse"] = f"{float(mse_val):.8f}"
            row["mae"] = f"{float(mae_val):.8f}"
            row["status"] = "ok"
        else:
            folder_path = exp._get_results_path(setting)
            mf = os.path.join(
                folder_path, f"{args.model}_{args.seq_len}_{args.pred_len}_metrics.npy"
            )
            if os.path.exists(mf):
                metrics = np.load(mf)
                mae_v, mse_v = float(metrics[0]), float(metrics[1])
                if np.isfinite(mse_v) and np.isfinite(mae_v):
                    row["mse"] = f"{mse_v:.8f}"
                    row["mae"] = f"{mae_v:.8f}"
                    row["status"] = "ok_fallback"
                else:
                    row["status"] = "nan_metrics"
            else:
                row["status"] = "no_metrics"

    except RuntimeError as e:
        if "out of memory" in str(e).lower():
            row["status"] = "oom"
            print(f"{tag} OOM")
        else:
            row["status"] = f"error:{type(e).__name__}"
            print(f"{tag} RuntimeError: {e}")
    except Exception as e:
        row["status"] = f"error:{type(e).__name__}"
        print(f"{tag} Exception: {e}")
        traceback.print_exc()
    finally:
        row["elapsed_s"] = f"{time.time() - t0:.1f}"
        if exp is not None:
            del exp
        gc.collect()
        torch.cuda.empty_cache()

    print(
        f"{tag} Done: status={row['status']}, mse={row.get('mse', 'N/A')}, "
        f"mae={row.get('mae', 'N/A')}, elapsed={row['elapsed_s']}s"
    )

    _write_xlsx_row(row)


if __name__ == "__main__":
    main()
